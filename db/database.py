"""Централізоване відкриття зашифрованої бази даних застосунку."""

from __future__ import annotations

from contextlib import contextmanager
from datetime import datetime
from pathlib import Path
import sqlite3 as plaintext_sqlite3
import threading

from openpyxl import Workbook, load_workbook
from sqlcipher3 import dbapi2 as sqlite3


DEFAULT_DB_PATH = Path(__file__).resolve().parent / "sodt.sqlite3"

_database_password: str | None = None
_thread_local = threading.local()


class DatabaseConfigurationError(RuntimeError):
    """Базовий виняток для помилок конфігурації доступу до БД."""


class DatabasePasswordError(DatabaseConfigurationError):
    """Помилка перевірки пароля або відкриття зашифрованої БД."""


class DatabaseMigrationError(DatabaseConfigurationError):
    """Помилка автоматичної міграції plaintext SQLite у зашифрований файл."""


class DatabaseRekeyError(DatabaseConfigurationError):
    """Помилка зміни пароля вже зашифрованої бази даних."""


class DatabaseTransferError(DatabaseConfigurationError):
    """Помилка експорту або імпорту даних бази даних."""


_EXCEL_META_SHEET = "__meta__"


def get_database_path(db_path: Path | None = None) -> Path:
    """Повертає канонічний шлях до файлу БД."""

    return Path(db_path) if db_path is not None else DEFAULT_DB_PATH


def get_database_state(db_path: Path | None = None) -> str:
    """Повертає стан файлу БД: missing, plaintext або encrypted."""

    resolved_path = get_database_path(db_path)
    if not resolved_path.exists():
        return "missing"
    if _is_plaintext_database(resolved_path):
        return "plaintext"
    return "encrypted"


def set_database_password(password: str) -> None:
    """Зберігає пароль поточного сеансу для всіх нових з'єднань."""

    global _database_password
    _database_password = _normalize_password(password)


def clear_database_password() -> None:
    """Скидає пароль поточного сеансу."""

    global _database_password
    _close_scoped_connections()
    _database_password = None


def is_database_configured() -> bool:
    """Ознака, що пароль для поточного процесу вже заданий."""

    return bool(_database_password)


def initialize_database_access(password: str, db_path: Path | None = None) -> Path:
    """Готує доступ до БД, а plaintext-файл мігрує у зашифрований формат."""

    resolved_path = get_database_path(db_path)
    normalized_password = _normalize_password(password)
    state = get_database_state(resolved_path)

    if state == "plaintext":
        _migrate_plaintext_database(resolved_path, normalized_password)

    set_database_password(normalized_password)
    try:
        with connection_scope(resolved_path):
            pass
    except sqlite3.DatabaseError as error:
        clear_database_password()
        raise DatabasePasswordError("Не вдалося відкрити зашифровану базу даних з указаним паролем.") from error

    return resolved_path


def change_database_password(current_password: str, new_password: str, db_path: Path | None = None) -> Path:
    """Перекодовує зашифровану БД на новий пароль."""

    resolved_path = get_database_path(db_path)
    state = get_database_state(resolved_path)
    if state == "missing":
        raise DatabaseRekeyError("Файл бази даних ще не створено.")
    if state == "plaintext":
        raise DatabaseRekeyError("Для plaintext-бази зміна пароля недоступна. Спочатку виконайте початкове шифрування.")

    normalized_current_password = _normalize_password(current_password)
    normalized_new_password = _normalize_password(new_password)
    previous_password = _database_password

    _close_scoped_connections()

    try:
        connection = _open_connection_with_password(resolved_path, normalized_current_password)
        try:
            connection.execute(f"PRAGMA rekey = '{_escape_pragma_value(normalized_new_password)}'")
            connection.commit()
        finally:
            connection.close()

        verification_connection = _open_connection_with_password(resolved_path, normalized_new_password)
        verification_connection.close()
    except sqlite3.DatabaseError as error:
        if previous_password:
            set_database_password(previous_password)
        else:
            clear_database_password()
        raise DatabaseRekeyError("Не вдалося перекодувати базу даних на новий пароль.") from error

    set_database_password(normalized_new_password)
    return resolved_path


def export_database_copy(destination_path: str | Path, db_path: Path | None = None) -> Path:
    """Експортує поточну базу даних у файл Excel."""

    resolved_path = get_database_path(db_path)
    destination = Path(destination_path)

    if not resolved_path.exists():
        raise DatabaseTransferError("Файл бази даних не знайдено.")

    destination.parent.mkdir(parents=True, exist_ok=True)

    try:
        workbook = Workbook()
        meta_sheet = workbook.active
        meta_sheet.title = _EXCEL_META_SHEET
        meta_sheet.append(["table_name", "sheet_name", "create_sql", "exported_at"])

        with connection_scope(resolved_path) as connection:
            table_rows = connection.execute(
                """
                SELECT name, sql
                FROM sqlite_master
                WHERE type = 'table'
                  AND name NOT LIKE 'sqlite_%'
                ORDER BY name
                """
            ).fetchall()

            used_sheet_names = {_EXCEL_META_SHEET.casefold()}
            exported_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            for table_row in table_rows:
                table_name = str(table_row["name"])
                create_sql = str(table_row["sql"] or "")
                sheet_name = _build_excel_sheet_name(table_name, used_sheet_names)
                meta_sheet.append([table_name, sheet_name, create_sql, exported_at])

                sheet = workbook.create_sheet(title=sheet_name)
                column_rows = connection.execute(f"PRAGMA table_info({_quote_identifier(table_name)})").fetchall()
                column_names = [str(column_row["name"]) for column_row in column_rows]
                sheet.append(column_names)

                if not column_names:
                    continue

                data_rows = connection.execute(f"SELECT * FROM {_quote_identifier(table_name)}").fetchall()
                for data_row in data_rows:
                    sheet.append([data_row[column_name] for column_name in column_names])

        workbook.save(destination)
    except (OSError, sqlite3.DatabaseError) as error:
        raise DatabaseTransferError("Не вдалося експортувати базу даних у файл Excel.") from error

    return destination


def import_database_copy(source_path: str | Path, db_path: Path | None = None) -> Path:
    """Імпортує дані з файла Excel у поточну базу даних."""

    if not _database_password:
        raise DatabaseTransferError("Імпорт доступний лише після відкриття бази даних з паролем.")

    resolved_path = get_database_path(db_path)
    source = Path(source_path)
    if not source.exists():
        raise DatabaseTransferError("Файл для імпорту не знайдено.")

    try:
        workbook = load_workbook(source)
    except OSError as error:
        raise DatabaseTransferError("Не вдалося відкрити файл Excel для імпорту.") from error
    except Exception as error:
        raise DatabaseTransferError(
            "Не вдалося прочитати файл Excel для імпорту. Переконайтеся, що це коректний файл експорту."
        ) from error

    if _EXCEL_META_SHEET not in workbook.sheetnames:
        raise DatabaseTransferError("У файлі Excel немає службового аркуша імпорту.")

    import_plan = _read_excel_import_plan(workbook)
    if not import_plan:
        raise DatabaseTransferError("У файлі Excel немає таблиць для імпорту.")

    try:
        _close_scoped_connections()

        with connection_scope(resolved_path) as connection:
            connection.execute("PRAGMA foreign_keys = OFF")
            current_tables = _get_user_table_names(connection)

            for table_name, create_sql, _sheet_name, _column_names, _rows in import_plan:
                if table_name not in current_tables and create_sql:
                    connection.execute(create_sql)

            current_tables = _get_user_table_names(connection)
            for table_name in current_tables:
                connection.execute(f"DELETE FROM {_quote_identifier(table_name)}")

            for table_name, _create_sql, _sheet_name, column_names, rows in import_plan:
                if not column_names:
                    continue
                normalized_rows = _normalize_excel_rows_for_table(connection, table_name, column_names, rows)
                placeholders = ", ".join("?" for _ in column_names)
                quoted_columns = ", ".join(_quote_identifier(column_name) for column_name in column_names)
                insert_sql = (
                    f"INSERT INTO {_quote_identifier(table_name)} ({quoted_columns}) VALUES ({placeholders})"
                )
                if normalized_rows:
                    connection.executemany(insert_sql, normalized_rows)

            connection.execute("PRAGMA foreign_keys = ON")
    except sqlite3.DatabaseError as error:
        raise DatabaseTransferError("Не вдалося імпортувати дані з файла Excel у базу даних.") from error

    return resolved_path


def _get_user_table_names(connection: sqlite3.Connection) -> list[str]:
    rows = connection.execute(
        """
        SELECT name
        FROM sqlite_master
        WHERE type = 'table'
          AND name NOT LIKE 'sqlite_%'
        ORDER BY name
        """
    ).fetchall()
    return [str(row["name"]) for row in rows]


def _build_excel_sheet_name(table_name: str, used_sheet_names: set[str]) -> str:
    base_name = table_name[:31] or "Sheet"
    candidate = base_name
    counter = 1
    while candidate.casefold() in used_sheet_names:
        suffix = f"_{counter}"
        candidate = f"{base_name[:31 - len(suffix)]}{suffix}"
        counter += 1
    used_sheet_names.add(candidate.casefold())
    return candidate


def _read_excel_import_plan(workbook) -> list[tuple[str, str, str, list[str], list[tuple[object, ...]]]]:
    meta_sheet = workbook[_EXCEL_META_SHEET]
    plan: list[tuple[str, str, str, list[str], list[tuple[object, ...]]]] = []

    for row in meta_sheet.iter_rows(min_row=2, values_only=True):
        table_name = str(row[0] or "").strip()
        sheet_name = str(row[1] or "").strip()
        create_sql = str(row[2] or "").strip()
        if not table_name or not sheet_name:
            continue
        if sheet_name not in workbook.sheetnames:
            raise DatabaseTransferError(f"У файлі Excel відсутній аркуш для таблиці «{table_name}».")

        sheet = workbook[sheet_name]
        raw_rows = list(sheet.iter_rows(values_only=True))
        if not raw_rows:
            plan.append((table_name, create_sql, sheet_name, [], []))
            continue

        column_names = [str(value or "").strip() for value in raw_rows[0] if str(value or "").strip()]
        data_rows = [
            tuple(_normalize_excel_cell_value(value) for value in data_row[: len(column_names)])
            for data_row in raw_rows[1:]
        ]
        plan.append((table_name, create_sql, sheet_name, column_names, data_rows))

    return plan


def _normalize_excel_cell_value(value: object) -> object:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    return value


def _normalize_excel_rows_for_table(
    connection: sqlite3.Connection,
    table_name: str,
    column_names: list[str],
    rows: list[tuple[object, ...]],
) -> list[tuple[object, ...]]:
    column_info_rows = connection.execute(f"PRAGMA table_info({_quote_identifier(table_name)})").fetchall()
    column_info_by_name = {
        str(column_info_row["name"]): column_info_row
        for column_info_row in column_info_rows
    }

    normalized_rows: list[tuple[object, ...]] = []
    for row in rows:
        normalized_values: list[object] = []
        for index, column_name in enumerate(column_names):
            value = row[index] if index < len(row) else None
            column_info = column_info_by_name.get(column_name)
            if value is None and column_info is not None and int(column_info["notnull"]) == 1:
                declared_type = str(column_info["type"] or "").upper()
                default_value = str(column_info["dflt_value"] or "")
                if "CHAR" in declared_type or "TEXT" in declared_type or "CLOB" in declared_type or default_value in ("''", '""'):
                    value = ""
            normalized_values.append(value)
        normalized_rows.append(tuple(normalized_values))

    return normalized_rows


def _quote_identifier(identifier: str) -> str:
    return '"' + identifier.replace('"', '""') + '"'


def connect(db_path: Path | None = None) -> sqlite3.Connection:
    """Створює SQLCipher-з'єднання з ключем поточного сеансу."""

    if not _database_password:
        raise DatabaseConfigurationError("Пароль для бази даних ще не налаштовано.")

    resolved_path = get_database_path(db_path)
    return _open_connection_with_password(resolved_path, _database_password)


def _open_connection_with_password(resolved_path: Path, password: str) -> sqlite3.Connection:
    """Відкриває нове валідоване SQLCipher-з'єднання з указаним паролем."""

    connection = sqlite3.connect(str(resolved_path))
    connection.row_factory = sqlite3.Row
    try:
        _apply_connection_settings(connection, password)
        connection.execute("SELECT count(*) FROM sqlite_master").fetchone()
    except sqlite3.DatabaseError:
        connection.close()
        raise
    return connection


@contextmanager
def connection_scope(db_path: Path | None = None):
    """Повертає з'єднання SQLCipher і гарантовано закриває його після використання."""

    connection_state = _get_scoped_connection_state(db_path)
    connection_state["depth"] += 1
    connection = connection_state["connection"]
    try:
        yield connection
        if connection_state["depth"] == 1 and connection.in_transaction:
            connection.commit()
    except Exception:
        if connection_state["depth"] == 1 and connection.in_transaction:
            connection.rollback()
        raise
    finally:
        connection_state["depth"] -= 1


def _apply_connection_settings(connection: sqlite3.Connection, password: str) -> None:
    connection.execute(f"PRAGMA key = '{_escape_pragma_value(password)}'")
    connection.execute("PRAGMA foreign_keys = ON")


def _get_scoped_connection_state(db_path: Path | None = None) -> dict[str, object]:
    """Повертає або створює кешований конекшн для поточного потоку."""

    cache = _get_scoped_connection_cache()
    resolved_path = get_database_path(db_path)
    cache_key = str(resolved_path.resolve())
    connection_state = cache.get(cache_key)
    if connection_state is not None and _is_connection_alive(connection_state["connection"]):
        return connection_state

    if connection_state is not None:
        try:
            connection_state["connection"].close()
        except sqlite3.DatabaseError:
            pass

    connection_state = {
        "connection": _open_connection_with_password(resolved_path, _database_password),
        "depth": 0,
    }
    cache[cache_key] = connection_state
    return connection_state


def _get_scoped_connection_cache() -> dict[str, dict[str, object]]:
    """Повертає кеш конекшнів для поточного потоку виконання."""

    cache = getattr(_thread_local, "scoped_connections", None)
    if cache is None:
        cache = {}
        _thread_local.scoped_connections = cache
    return cache


def _is_connection_alive(connection: sqlite3.Connection) -> bool:
    """Швидко перевіряє, чи кешоване з'єднання ще придатне до роботи."""

    try:
        connection.execute("SELECT 1").fetchone()
    except sqlite3.DatabaseError:
        return False
    return True


def _close_scoped_connections() -> None:
    """Закриває всі кешовані з'єднання поточного потоку."""

    cache = _get_scoped_connection_cache()
    for connection_state in cache.values():
        connection = connection_state["connection"]
        try:
            connection.close()
        except sqlite3.DatabaseError:
            pass
    cache.clear()


def _normalize_password(password: str) -> str:
    normalized_password = password.strip()
    if not normalized_password:
        raise DatabasePasswordError("Пароль до бази даних не може бути порожнім.")
    return normalized_password


def _escape_pragma_value(value: str) -> str:
    return value.replace("'", "''")


def _is_plaintext_database(db_path: Path) -> bool:
    connection = None
    try:
        connection = plaintext_sqlite3.connect(db_path)
        connection.execute("SELECT count(*) FROM sqlite_master").fetchone()
    except plaintext_sqlite3.DatabaseError:
        return False
    finally:
        if connection is not None:
            connection.close()
    return True


def _migrate_plaintext_database(db_path: Path, password: str) -> None:
    """Копіює наявну plaintext SQLite-базу в новий SQLCipher-файл."""

    temporary_path = db_path.with_name(f"{db_path.stem}.encrypted{db_path.suffix}")
    backup_path = db_path.with_name(f"{db_path.stem}.plaintext-backup{db_path.suffix}")

    if temporary_path.exists():
        temporary_path.unlink()

    try:
        source_connection = plaintext_sqlite3.connect(db_path)
        try:
            dump_script = "\n".join(list(source_connection.iterdump()))
        finally:
            source_connection.close()

        target_connection = sqlite3.connect(str(temporary_path))
        try:
            _apply_connection_settings(target_connection, password)
            target_connection.execute("PRAGMA foreign_keys = OFF")
            target_connection.executescript(dump_script)
            target_connection.commit()
            target_connection.execute("PRAGMA foreign_keys = ON")
            target_connection.execute("SELECT count(*) FROM sqlite_master").fetchone()
        finally:
            target_connection.close()
    except (plaintext_sqlite3.DatabaseError, sqlite3.DatabaseError, OSError) as error:
        if temporary_path.exists():
            temporary_path.unlink()
        raise DatabaseMigrationError("Не вдалося мігрувати наявну plaintext-базу у зашифрований формат.") from error

    if backup_path.exists():
        backup_path.unlink()

    db_path.replace(backup_path)
    temporary_path.replace(db_path)